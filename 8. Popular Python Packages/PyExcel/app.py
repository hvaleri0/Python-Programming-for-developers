import openpyxl

# wb = openpyxl.workbook()  # create new workbook
wb = openpyxl.load_workbook("transactions.xlsx")  # open existing workbook
print(wb.sheetnames)  # gets you the name of sheets in the workbook

# to access the sheet in the work book, note tath the string is case sensitive:
sheet = wb["Sheet1"]

# You can create a new sheet or remove exxisting ones:
# wb.create_sheet("Sheet2", 0)  # index zero means to out it before sheet1
# wb.remove_sheet(sheet) # to remove the sheet

# different ways to access individual or range of cells:
cell = sheet["a1"]
cell.value = "change the value"
print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)

# usefull when iterating over evey row and column
cell = sheet.cell(row=1, column=1)
print(sheet.max_row)
print(sheet.max_column)

# iterate:
for row in range(1, sheet.max_row+1):  # there is no row zero thats why we start from 1
    for column in range(1, sheet.max_column+1):
        cell = sheet.cell(row, column)
        print(cell.value)

column = sheet['a']  # return all the cells in column a
print(column)

cells = sheet['a:c']  # return all cells from a to c
print(cells)

cells2 = sheet['a1:c3']  # return all cells from a1 to c3 coodinates
print(cells2)

rows = sheet[1:3]  # returns by row eg row 1 through 3
print(rows)


# How to insert/add/delete:
sheet.append([1, 2, 3])  # add a row at the end of the sheet:
# sheet.insert_cols
# sheet.insert_rows
# sheet.delete_cols
# sheet.delete_rows
wb.save("transactions2.xlsx")  # create a nd save to new workbook
